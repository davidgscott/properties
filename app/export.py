"""Export screened results to .xlsx or .csv."""
from __future__ import annotations

import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

COLUMNS = [
    ("apn", "APN"),
    ("address", "Address"),
    ("jurisdiction", "Jurisdiction"),
    ("zoning", "Zoning"),
    ("permit_type", "Storage permit"),
    ("zoning_confidence", "Zoning conf."),
    ("flood_zone", "FEMA zone"),
    ("sfha_pct", "% in SFHA"),
    ("floodway", "Floodway"),
    ("slope_mean_pct", "Slope mean %"),
    ("slope_max_pct", "Slope max %"),
    ("acres", "Acres"),
    ("vacant_category", "Vacant/Use"),
    ("listed", "Listed"),
    ("list_price", "List price"),
    ("status", "Status"),
    ("score", "Score"),
    ("reasons_text", "Notes / reasons"),
]

_STATUS_FILL = {
    "PASS": PatternFill("solid", fgColor="C6EFCE"),
    "REVIEW": PatternFill("solid", fgColor="FFEB9C"),
    "FAIL": PatternFill("solid", fgColor="FFC7CE"),
}


def _row_values(r: dict) -> dict:
    r = dict(r)
    r["reasons_text"] = "; ".join(r.get("reasons", []))
    r["floodway"] = "YES" if r.get("floodway") else ""
    r["listed"] = "YES" if r.get("listed") else ""
    return r


def to_csv(results: list[dict]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([h for _, h in COLUMNS])
    for r in results:
        rv = _row_values(r)
        w.writerow([rv.get(k, "") for k, _ in COLUMNS])
    return buf.getvalue().encode("utf-8")


def to_xlsx(results: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Storage screen"
    ws.append([h for _, h in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    status_col = [k for k, _ in COLUMNS].index("status") + 1
    for r in results:
        rv = _row_values(r)
        ws.append([rv.get(k, "") for k, _ in COLUMNS])
        fill = _STATUS_FILL.get(rv.get("status"))
        if fill:
            ws.cell(row=ws.max_row, column=status_col).fill = fill
    # Reasonable column widths.
    for i, (_, header) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = \
            max(10, min(48, len(header) + 4))
    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
